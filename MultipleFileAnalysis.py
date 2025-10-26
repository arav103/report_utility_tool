import pandas as pd
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import os
from Utils import add_campaign_details_rows, extract_campaign_details, parse_html


def extract_messages_from_files(filepaths):
    """
    Extract messages from multiple files and organize them by test case and date.

    """
    all_results = defaultdict(
        lambda: defaultdict(
            lambda: {"Pass": 0, "Fail": 0, "Error": 0, "Warning": 0, "Total": 0}
        )
    )
    campaign_details = {}
    dates = []


    for filepath in filepaths:
        soup = parse_html(filepath)

        campaign_date, details = extract_campaign_details(filepath, soup)
        campaign_details[filepath] = details
        dates.append(campaign_date)

        for test_div in soup.find_all("div", attrs={"name": "test"}):
            name_tag = test_div.find_next("b", text="Name")
            valuation_tag = test_div.find_next("b", text="Valuation")
            if name_tag and valuation_tag:
                test_case_name = (
                    name_tag.find_next_sibling(text=True)
                    .strip()
                    .replace(":", "")
                    .strip()
                )
                valuation = (
                    valuation_tag.find_next_sibling(text=True)
                    .strip()
                    .replace(":", "")
                    .strip()
                    .upper()
                )
                if test_case_name:
                    all_results[test_case_name][campaign_date]["Total"] += 1
                    if valuation == "PASS":
                        all_results[test_case_name][campaign_date]["Pass"] += 1
                    elif valuation == "FAIL":
                        all_results[test_case_name][campaign_date]["Fail"] += 1
                    elif valuation == "ERROR":
                        all_results[test_case_name][campaign_date]["Error"] += 1
                    elif valuation == "WARNING":
                        all_results[test_case_name][campaign_date]["Warning"] += 1

    return all_results, sorted(set(dates)), campaign_details


def prepare_details_sheet_data(results, dates):
    """
    Prepare data for the "Details" sheet in the report.

    """
    test_case_data = []
    date_columns = {date: [] for date in dates}

    for test_case, values in results.items():
        row = {"Test Case": test_case}
        total_runs, total_pass, total_fail, total_error, total_warning = 0, 0, 0, 0, 0

        for date in dates:
            if date in values and values[date]["Total"] > 0:
                row[date] = f"{values[date]['Pass']}/{values[date]['Total']}"
                date_columns[date].append(values[date])
                total_pass += values[date]["Pass"]
                total_fail += values[date]["Fail"]
                total_error += values[date]["Error"]
                total_warning += values[date]["Warning"]
                total_runs += values[date]["Total"]
            else:
                row[date] = "--"

        row["Total Runs"] = total_runs
        row["Passes"] = total_pass
        row["Fails"] = total_fail
        row["Errors"] = total_error
        row["Warnings"] = total_warning
        row["Stability (%)"] = (
            f"{(total_pass / total_runs * 100):.2f}%" if total_runs > 0 else "--"
        )
        test_case_data.append(row)

    return pd.DataFrame(test_case_data), date_columns

def generate_summary_plot(date_columns, dates):
    """
    Generate a stacked bar chart for test case counts by date.

    """
    plot_data = []
    for date in dates:
        date_data = defaultdict(int)
        for counts in date_columns[date]:
            for key in ["Pass", "Fail", "Error", "Warning"]:
                date_data[key] += counts[key]
        date_data["Date"] = date
        plot_data.append(date_data)

    plot_df = pd.DataFrame(plot_data)

    fig, ax = plt.subplots(figsize=(12, 6))
    cumulative_bottom = pd.Series([0] * len(plot_df))

    for category, color in zip(
        ["Pass", "Fail", "Error", "Warning"], ["green", "red", "orange", "yellow"]
    ):
        bars = ax.bar(
            plot_df["Date"],
            plot_df[category],
            color=color,
            label=category,
            bottom=cumulative_bottom,
        )
        for bar in bars:
            if bar.get_height() > 0:
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_y() + bar.get_height() / 2,
                    int(bar.get_height()),
                    ha="center",
                    va="center",
                    fontsize=9,
                    color="white",
                )
        cumulative_bottom += plot_df[category]

    ax.legend()
    ax.set_title("Test Case Counts by Date")
    ax.set_xlabel("Dates")
    ax.set_ylabel("Counts")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig("summary_plot.png")
    plt.close()

def generate_cyclic_summary_plot(results, dates, output_path, worksheet):
    """
    Generate a cyclic summary bar plot and add details to the worksheet.

    """
    cyclic_data = defaultdict(lambda: {"Pass": 0, "Fail": 0})
    passed_failed_details = defaultdict(lambda: {"Pass": [], "Fail": []})

    for test_case, date_results in results.items():
        for date in dates:
            if date in date_results:
                counts = date_results[date]
                if counts["Fail"] > 0 or counts["Error"] > 0:
                    cyclic_data[date]["Fail"] += 1
                    passed_failed_details[date]["Fail"].append(test_case)
                elif counts["Pass"] > 0:
                    cyclic_data[date]["Pass"] += 1
                    passed_failed_details[date]["Pass"].append(test_case)

    cyclic_df = pd.DataFrame.from_dict(cyclic_data, orient="index").reset_index()
    cyclic_df.columns = ["Date", "Pass", "Fail"]

    fig, ax = plt.subplots(figsize=(10, 6))
    x = range(len(cyclic_df["Date"]))
    cumulative_bottom = [0] * len(cyclic_df)

    for category, color in zip(["Pass", "Fail"], ["green", "red"]):
        heights = cyclic_df[category]
        bars = ax.bar(x, heights, color=color, bottom=cumulative_bottom, label=category)

        for i, bar in enumerate(bars):
            if bar.get_height() > 0:
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    cumulative_bottom[i] + bar.get_height() / 2,
                    int(bar.get_height()),
                    ha="center",
                    va="center",
                    fontsize=9,
                    color="white",
                )
        cumulative_bottom = [sum(val) for val in zip(cumulative_bottom, heights)]

    ax.set_xticks(x)
    ax.set_xticklabels(cyclic_df["Date"], rotation=45, ha="right")
    ax.set_xlabel("Date")
    ax.set_ylabel("Count")
    ax.set_title("Cyclic Summary by Date")
    ax.legend()
    plt.tight_layout()
    plt.savefig(output_path)
    plt.close()

    row_idx = 36
    for date in dates:
        worksheet.cell(row=row_idx, column=1, value=f"Date: {date}").font = Font(bold=True)
        row_idx += 1

        worksheet.cell(row=row_idx, column=1, value="Passed Scripts:").font = Font(bold=True)
        worksheet.cell(row=row_idx, column=2, value=", ".join(passed_failed_details[date]["Pass"]))
        row_idx += 1

        worksheet.cell(row=row_idx, column=1, value="Failed Scripts:").font = Font(bold=True)
        worksheet.cell(row=row_idx, column=2, value=", ".join(passed_failed_details[date]["Fail"]))
        row_idx += 2

def generate_multi_file_summary(filepaths, save_path):
    """
    Generate a summary report for multiple files.

    """
    results, dates, campaign_details = extract_messages_from_files(filepaths)
    details_df, date_columns = prepare_details_sheet_data(results, dates)
    add_campaign_details_rows(details_df, campaign_details, dates)

    output_file = os.path.join(save_path, "MultiFileAnalysis_Summary.xlsx")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        details_df.to_excel(writer, sheet_name="Details", index=False)

        generate_summary_plot(date_columns, dates)
        summary_plot_sheet = writer.book.create_sheet("Summary Plot")
        summary_img = Image("summary_plot.png")
        summary_plot_sheet.add_image(summary_img, "A1")

        cyclic_plot_path = "cyclic_summary_plot.png"
        cyclic_plot_sheet = writer.book.create_sheet("Cyclic Summary Plot")
        generate_cyclic_summary_plot(results, dates, cyclic_plot_path, cyclic_plot_sheet)
        cyclic_img = Image(cyclic_plot_path)
        cyclic_plot_sheet.add_image(cyclic_img, "A1")

    print(f"Summary report saved to {output_file}")
