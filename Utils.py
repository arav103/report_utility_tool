from bs4 import BeautifulSoup
import re
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from collections import defaultdict


def parse_html(filepath):
    """
    Parse the HTML content from a file.

    """
    with open(filepath, "r", encoding="utf-8") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")
    return soup


def add_campaign_details_rows(df, campaign_details, dates):
    """
    Add campaign details rows to a DataFrame for reporting.

    """
    campaign_row = {"Test Case": "Campaign Details", **{date: "" for date in dates}}
    bench_row = {"Test Case": "Bench Name", **{date: "" for date in dates}}
    python_version_row = {"Test Case": "Python Version", **{date: "" for date in dates}}
    enna_version_row = {"Test Case": "ENNA Version", **{date: "" for date in dates}}
    train_row = {"Test Case": "Train", **{date: "" for date in dates}}

    for filepath, details in campaign_details.items():
        for date in dates:
            if date in filepath:
                bench_row[date] = details.get("Testbench", "N/A")
                python_version_row[date] = details.get("Python Version", "N/A")
                enna_version_row[date] = details.get("ENNA Version", "N/A")
                train_row[date] = details.get("Train", "N/A")

    df.loc[len(df)] = campaign_row
    df.loc[len(df)] = bench_row
    df.loc[len(df)] = python_version_row
    df.loc[len(df)] = enna_version_row
    df.loc[len(df)] = train_row


def extract_campaign_details(filepath, soup):
    """
    Extract campaign details from the parsed HTML content.

    """
    campaign_date = None
    testbench = filepath.split("/")[-1][:5]
    campaign_details = {"Testbench": testbench,
                        "Python Version": "Unknown", "ENNA Version": "Unknown", "Train": "Unknown"}

    campaign_section = soup.find("div", class_="content active")
    if campaign_section:
        table_rows = campaign_section.find_all("tr")
        for row in table_rows:
            columns = row.find_all("td")
            if len(columns) == 2:
                key = columns[0].get_text(strip=True)
                value = columns[1].get_text(strip=True)
                if key == "Campaign date":
                    campaign_date = value.split()[0]
                elif key == "ENNA version":
                    campaign_details["ENNA Version"] = value
                elif key == "Python version":
                    campaign_details["Python Version"] = value
                elif key == "Train":
                    campaign_details["Train"] = value

    if not campaign_date:
        campaign_date = "Unknown Date"

    return campaign_date, campaign_details


def find_closest_test_case(test_cases, reference_line):
    """
    Find the closest test case based on the reference line in the HTML.

    """
    closest_test_case = "Unknown Test Case"
    min_distance = float("inf")
    for test_div, test_id in test_cases:
        distance = abs(test_div.sourceline - reference_line)
        if distance < min_distance:
            closest_test_case = test_id
            min_distance = distance
    return closest_test_case


def find_closest_stimulation(stimulations, reference_line):
    """
     Find the closest stimulation based on the reference line in the HTML.

     """
    for stim_div, stim_name in reversed(stimulations):
        if stim_div.sourceline < reference_line:
            return stim_name
    return "Unknown Stimulation"


def extract_previous_actions(tag):
    """
    Extract previous actions from the HTML tag.

    """
    previous_actions = []
    for sibling in tag.find_previous_siblings():
        if sibling.name == "span" and "text-info" in sibling.get("class", []):
            previous_actions.append(sibling.get_text(strip=True))
        if len(previous_actions) >= 3:
            break
    return "; ".join(previous_actions[::-1])


def parse_issues(tag, stimulations):
    """
    Parse issue details (e.g., errors or failures) from an HTML tag.

    """
    class_name = tag["class"][0]
    if class_name not in ["text-error", "text-fail"]:
        return None

    text = tag.get_text(strip=True)
    parts = text.split('|')
    if len(parts) >= 4:
        message = parts[3].strip()
        timestamp = re.sub(r"[ a-zA-Z]", "", parts[0].strip())
    elif len(parts) >= 3:
        message = parts[2].strip()
        timestamp = re.sub(r"[ a-zA-Z]", "", parts[0].strip())
    else:
        return None

    if len(message) > 200:
        for sibling in tag.find_next_siblings("span"):
            if sibling["class"][0] == "text-error":
                message += f" {sibling.get_text(strip=True)}"
            else:
                break

    test_case_match = re.search(r"(\d{2,}_[A-Za-z0-9_]+)", message)
    if not test_case_match:
        return None
    test_case_name = test_case_match.group(1)

    stimulation = find_closest_stimulation(stimulations, tag.sourceline)
    previous_actions = extract_previous_actions(tag)

    return {
        "stimulation": stimulation,
        "test_case": test_case_name,
        "message": message,
        "type": "Error" if class_name == "text-error" else "Failure",
        "timestamp": timestamp,
        "previous_actions": previous_actions,
    }


def process_content(soup, stimulations, test_cases, keyword, target_dict):
    for div in soup.find_all("div", class_="content"):
        if "Valuation" in div.text and keyword in div.text:
            stimulation = find_closest_stimulation(stimulations, div.sourceline)
            test_case = find_closest_test_case(test_cases, div.sourceline)
            # Always add to target_dict, regardless of duplicates
            target_dict[stimulation]["stimulations"].append(stimulation)
            target_dict[stimulation]["test_cases"].append(test_case)


def process_nonduplicate(soup, stimulations, test_cases, keyword, target_dict):
    for div in soup.find_all("div", class_="content"):
        if "Valuation" in div.text and keyword in div.text:
            stimulation = find_closest_stimulation(stimulations, div.sourceline)

            test_case = find_closest_test_case(test_cases, div.sourceline)

            if stimulation not in target_dict or test_case not in target_dict[stimulation]["test_cases"]:
                target_dict[stimulation]["stimulations"].append(stimulation)
                target_dict[stimulation]["test_cases"].append(test_case)

    for stim_key, stim_data in target_dict.items():
        stim_data["stimulations"] = list(set(stim_data["stimulations"]))
        stim_data["test_cases"] = list(set(stim_data["test_cases"]))


def prepare_message_rows(data, is_pass=False, retain_duplicates=False):
    rows = []
    seen_entries = set()
    for stim, info in data.items():
        for i in range(len(info["test_cases"])):
            test_case = info["test_cases"][i]
            stimulation = stim
            if is_pass:
                row = (test_case, stimulation, "")
            else:
                row = (
                    test_case,
                    info["types"][i],
                    stimulation,
                    info["messages"][i] if "messages" in info else "",
                    info["times"][i],
                    info["previous_actions"][i],
                )

            if retain_duplicates or row not in seen_entries:
                rows.append(row)
                seen_entries.add(row)
    return rows


def generate_summary_piechart(writer, category_counts, sheet_name="Summary"):
    """
    Generate a pie chart and write it to the summary sheet.

    """
    categories = list(category_counts.keys())
    counts = list(category_counts.values())
    colors = ["green", "yellow", "orange", "red"]

    plt.figure(figsize=(5, 5))
    plt.pie(counts, labels=categories, colors=colors, autopct='%1.1f%%', startangle=140)
    plt.title("Summary")
    plt.savefig("pie_chart.png")
    plt.close()

    summary_sheet = writer.book.create_sheet(sheet_name)
    img = Image("pie_chart.png")
    summary_sheet.add_image(img, "A1")

    for idx, (category, count) in enumerate(category_counts.items(), start=2):
        summary_sheet.cell(row=idx, column=1, value=category).font = Font(bold=True)
        summary_sheet.cell(row=idx, column=2, value=count)


def generate_excel_report(output_file, passes_df, warnings_df, issues_df, campaign_details):
    """
    Generate an Excel report with issues, warnings, passes, and campaign details.

    """
    category_counts = {
        "Passes": len(passes_df),
        "Warnings": len(warnings_df),
        "Failures": len(issues_df[issues_df["Type"] == "Failure"]),
        "Errors": len(issues_df[issues_df["Type"] == "Error"]),
    }

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        issues_df.to_excel(writer, sheet_name="Issues", index=False)
        warnings_df.to_excel(writer, sheet_name="Warnings", index=False)
        passes_df.to_excel(writer, sheet_name="Passes", index=False)

        summary_sheet = writer.book.create_sheet("Summary")
        generate_summary_piechart(writer, category_counts, sheet_name="Summary")

        for idx, (category, count) in enumerate(category_counts.items(), start=2):
            summary_sheet.cell(row=idx, column=1, value=category).font = Font(bold=True)
            summary_sheet.cell(row=idx, column=2, value=count)

        start_row = len(category_counts) + 4
        for idx, (key, value) in enumerate(campaign_details.items(), start=start_row):
            summary_sheet.cell(row=idx, column=1, value=key).font = Font(bold=True)
            summary_sheet.cell(row=idx, column=2, value=value)

    print(f"Excel report with issues, warnings, passes, and campaign details written to {output_file}")
