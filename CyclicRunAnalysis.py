import pandas as pd
from collections import defaultdict
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from Utils import (
    parse_html,
    prepare_message_rows,
    generate_summary_piechart,
    parse_issues,
    find_closest_stimulation,
    find_closest_test_case
)

def process_content_cyclic_run(soup, stimulations, test_cases, keyword, target_dict):
    """
    Process HTML content for cyclic runs, allowing duplicates for passes and warnings.

    """
    for div in soup.find_all("div", class_="content"):
        if "Valuation" in div.text and keyword in div.text:
            stimulation = find_closest_stimulation(stimulations, div.sourceline)
            test_case = find_closest_test_case(test_cases, div.sourceline)
            target_dict[stimulation]["stimulations"].append(stimulation)
            target_dict[stimulation]["test_cases"].append(test_case)

def extract_cyclic_messages(html_file):
    """
    Extract messages from the provided HTML file for cyclic run analysis.

    """
    soup = parse_html(html_file)

    passes = defaultdict(lambda: {"stimulations": [], "test_cases": []})
    warnings = defaultdict(lambda: {"stimulations": [], "test_cases": []})
    issues = defaultdict(lambda: {
        "stimulations": [],
        "test_cases": [],
        "messages": [],
        "types": [],
        "times": [],
        "previous_actions": []
    })

    stimulations = [
        (div, div.find("span", class_="highlight").get_text(strip=True))
        for div in soup.find_all("div") if "title" in div.get("class", []) and div.find("span", class_="highlight")
    ]

    test_cases = [
        (div, div.get_text(strip=True).split()[0])
        for div in soup.find_all("div") if "title" in div.get("class", []) and "test" in div.get("class", [])
    ]

    for tag in soup.find_all("span", class_=True):
        issue = parse_issues(tag, stimulations)
        if issue:
            issues[issue["stimulation"]]["stimulations"].append(issue["stimulation"])
            issues[issue["stimulation"]]["test_cases"].append(issue["test_case"])
            issues[issue["stimulation"]]["messages"].append(issue["message"])
            issues[issue["stimulation"]]["types"].append(issue["type"])
            issues[issue["stimulation"]]["times"].append(issue["timestamp"])
            issues[issue["stimulation"]]["previous_actions"].append(issue["previous_actions"])

    process_content_cyclic_run(soup, stimulations, test_cases, "PASS", passes)

    process_content_cyclic_run(soup, stimulations, test_cases, "WARNING", warnings)

    campaign_details = {}
    campaign_section = soup.find("div", {"data-tab": "campaign"})
    if campaign_section:
        for row in campaign_section.find_all("tr"):
            columns = row.find_all("td")
            if len(columns) == 2:
                key = columns[0].get_text(strip=True)
                value = columns[1].get_text(strip=True)
                campaign_details[key] = value
    return passes, issues, warnings, campaign_details


def generate_excel_report(output_file, passes, warnings, issues, campaign_details):
    """
    Generate an Excel report with issues, warnings, passes, and campaign details.

    """
    issues_data_frame = pd.DataFrame(
        prepare_message_rows(issues),
        columns=["Test Case", "Type", "Stimulation", "Message", "Time", "Previous Actions"],
    )
    issues_data_frame.insert(5, "Jira Ticket", "")
    issues_data_frame.insert(6, "Jira Status", "")
    issues_data_frame.insert(7, "Comments", "")

    passes_data_frame = pd.DataFrame(
        prepare_message_rows(passes, is_pass=True, retain_duplicates=True),
        columns=["Test Case", "Stimulation", "Message"],
    )

    warnings_data_frame = pd.DataFrame(
        prepare_message_rows(warnings, is_pass=True, retain_duplicates=True),
        columns=["Test Case", "Stimulation", "Message"],
    )

    if not passes_data_frame.empty:
        passes_data_frame = passes_data_frame.iloc[1:]
    if not warnings_data_frame.empty:
        warnings_data_frame = warnings_data_frame.iloc[1:]

    total_passes = (len(passes_data_frame) - 1) // 2 if not passes_data_frame.empty else 0
    total_warnings = (len(warnings_data_frame) - 1) // 2 if not warnings_data_frame.empty else 0
    total_failures = len(issues_data_frame[issues_data_frame["Type"] == "Failure"])
    total_errors = len(issues_data_frame[issues_data_frame["Type"] == "Error"])
    category_counts = {
        "Passes": total_passes,
        "Warnings": total_warnings,
        "Failures": total_failures,
        "Errors": total_errors,
    }

    unique_passes_df = passes_data_frame.drop_duplicates(subset=["Test Case", "Stimulation"]).reset_index(drop=True)
    unique_warnings_df = warnings_data_frame.drop_duplicates(subset=["Test Case", "Stimulation"]).reset_index(drop=True)

    for df in [issues_data_frame, unique_passes_df, unique_warnings_df]:
        df["Test Case"] = df["Test Case"].mask(df["Test Case"].duplicated(), "")

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        issues_data_frame.to_excel(writer, sheet_name="Issues", index=False)
        unique_warnings_df.to_excel(writer, sheet_name="Warnings", index=False)
        unique_passes_df.to_excel(writer, sheet_name="Passes", index=False)

        generate_summary_piechart(writer, category_counts)

        workbook = writer.book
        summary_sheet = workbook["Summary"]

        campaign_info = [
            ("Campaign Name", "Campaign name"),
            ("Campaign Date", "Campaign date"),
            ("Duration", "Duration"),
            ("ENNA Version", "ENNA version"),
            ("Python Version", "Python version"),
            ("Train", "Train"),
        ]

        row_idx = 1
        for label, key in campaign_info:
            if key in campaign_details:
                summary_sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                summary_sheet.cell(row=row_idx, column=2, value=campaign_details[key])
                row_idx += 1

        start_row = row_idx + 2
        for idx, (category, count) in enumerate(category_counts.items(), start=start_row):
            summary_sheet.cell(row=idx, column=1, value=category).font = Font(bold=True)
            summary_sheet.cell(row=idx, column=2, value=count)

        pie_chart_image = "summary_piechart.png"
        img = Image(pie_chart_image)
        summary_sheet.add_image(img, "E1")

    print(f"Excel report generated at {output_file}")


def analyze_cyclic_run(html_file, save_path):
    """
    Analyze the provided HTML file for cyclic run data and generate a summary report.

    """
    passes, issues, warnings, campaign_details = extract_cyclic_messages(html_file)

    output_file = f"{save_path}/{html_file.split('/')[-1]}_cyclic_summary.xlsx"

    generate_excel_report(output_file, passes, warnings, issues, campaign_details)

